import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, NumberFormatDescriptor
import datetime

def generate_thesis_excel(
    baseline_price: float,
    strike_price: float,
    entry_premium: float,
    multiplier: int,
    iv: float,
    risk_free_rate: float,
    expiration_date: datetime.date,
    starting_dte: int,
    option_type: str = "PUT"
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    # --- 1. Assumptions Sheet ---
    ws_assumptions = wb.create_sheet("Assumptions")
    ws_assumptions.column_dimensions['A'].width = 30
    ws_assumptions.column_dimensions['B'].width = 15
    ws_assumptions.column_dimensions['C'].width = 15
    ws_assumptions.column_dimensions['D'].width = 60
    
    ws_assumptions.cell(row=1, column=1, value="SPY Options Scenario Model").font = Font(bold=True)
    
    headers = ["Input", "Value", "Units", "Notes / Source"]
    for i, h in enumerate(headers, 1):
        c = ws_assumptions.cell(row=3, column=i, value=h)
        c.font = Font(bold=True)
    
    inputs = [
        ("Baseline SPY price", baseline_price, "$/share", "Current underlying price"),
        ("Strike", strike_price, "$/share", f"Selected {option_type} strike"),
        ("Entry premium", entry_premium, "$/share", "Limit entry price"),
        ("Contract multiplier", multiplier, "shares/contract", "Standard multiplier"),
        ("Implied volatility", iv, "%", "IV calibrated to current mark"),
        ("Risk-free rate", risk_free_rate, "%", "Short-term risk-free rate"),
        ("Expiration date", expiration_date, "date", "Contract expiration"),
        ("Starting DTE", starting_dte, "calendar days", "Days to expiration")
    ]
    
    for row_idx, (name, val, unit, note) in enumerate(inputs, start=4):
        ws_assumptions.cell(row=row_idx, column=1, value=name)
        val_cell = ws_assumptions.cell(row=row_idx, column=2, value=val)
        if name in ["Baseline SPY price", "Strike", "Entry premium"]:
            val_cell.number_format = '$#,##0.00'
        elif name in ["Implied volatility", "Risk-free rate"]:
            val_cell.number_format = '0.00%'
        elif name == "Expiration date":
            val_cell.number_format = 'yyyy-mm-dd'
        ws_assumptions.cell(row=row_idx, column=3, value=unit)
        ws_assumptions.cell(row=row_idx, column=4, value=note)

    ws_assumptions.cell(row=13, column=1, value="Key outputs").font = Font(bold=True)
    ws_assumptions.cell(row=14, column=1, value="Total premium at risk")
    ws_assumptions.cell(row=14, column=2, value=f"=B6*B7").number_format = '$#,##0.00'
    ws_assumptions.cell(row=15, column=1, value="Expiration breakeven")
    if option_type.upper() == "PUT":
        ws_assumptions.cell(row=15, column=2, value=f"=B5-B6").number_format = '$#,##0.00'
    else:
        ws_assumptions.cell(row=15, column=2, value=f"=B5+B6").number_format = '$#,##0.00'
        
    ws_assumptions.cell(row=16, column=1, value="Breakeven % from baseline")
    if option_type.upper() == "PUT":
        ws_assumptions.cell(row=16, column=2, value=f"=1-(B15/B4)").number_format = '0.00%'
    else:
        ws_assumptions.cell(row=16, column=2, value=f"=(B15/B4)-1").number_format = '0.00%'
        
    # Generate Date Columns (0 to starting_dte)
    # We want starting_dte columns + 1 (day 0)
    # E.g., if DTE=3, dates are today, tomorrow, day after, exp day
    # So headers go from left to right: starting_dte down to 0
    date_columns = []
    base_date = expiration_date - datetime.timedelta(days=starting_dte)
    for dte_val in range(starting_dte, -1, -1):
        cur_date = expiration_date - datetime.timedelta(days=dte_val)
        date_columns.append((dte_val, cur_date))

    # Calculate percentage steps (+10% to -10%)
    pct_steps = []
    for i in range(10, -11, -1):
        pct_steps.append(i / 100.0)

    # --- 2. Option Price Matrix ---
    ws_price = wb.create_sheet("Option Price Matrix")
    ws_price.cell(row=1, column=1, value="Option Price Matrix").font = Font(bold=True)
    ws_price.cell(row=2, column=1, value="Rows = Underlying % change. Columns = calendar dates remaining.")
    
    ws_price.cell(row=3, column=1, value="% Change").font = Font(bold=True)
    ws_price.cell(row=3, column=2, value="Price").font = Font(bold=True)
    
    for col_idx, (dte, dt) in enumerate(date_columns, start=3):
        c = ws_price.cell(row=3, column=col_idx, value=dt)
        c.number_format = 'mmm d, yyyy'
        c.font = Font(bold=True)
        # Store DTE in row 24 for formulas
        ws_price.cell(row=24, column=col_idx, value=dte)

    for r_idx, pct in enumerate(pct_steps, start=4):
        ws_price.cell(row=r_idx, column=1, value=pct).number_format = '0.00%'
        ws_price.cell(row=r_idx, column=2, value=f"='Assumptions'!$B$4*(1+A{r_idx})").number_format = '$#,##0.00'
        
        for col_idx, (dte, dt) in enumerate(date_columns, start=3):
            col_letter = get_column_letter(col_idx)
            # Black-Scholes Formula in Excel
            S = f"$B{r_idx}"
            K = "'Assumptions'!$B$5"
            T = f"({col_letter}$24/365)"
            r = "'Assumptions'!$B$9"
            v = "'Assumptions'!$B$8"
            
            d1 = f"((LN({S}/{K})+({r}+0.5*{v}^2)*{T})/({v}*SQRT({T})))"
            d2 = f"({d1}-{v}*SQRT({T}))"
            
            if option_type.upper() == "CALL":
                # Call Price = S * N(d1) - K * exp(-rT) * N(d2)
                bs_formula = f"IF({col_letter}$24=0,MAX({S}-{K},0),{S}*NORMSDIST({d1})-{K}*EXP(-{r}*{T})*NORMSDIST({d2}))"
            else:
                # Put Price = K * exp(-rT) * N(-d2) - S * N(-d1)
                bs_formula = f"IF({col_letter}$24=0,MAX({K}-{S},0),{K}*EXP(-{r}*{T})*NORMSDIST(-{d2})-{S}*NORMSDIST(-{d1}))"
            
            ws_price.cell(row=r_idx, column=col_idx, value=f"={bs_formula}").number_format = '$#,##0.00'

    # --- 3. P&L Matrix ---
    ws_pnl = wb.create_sheet("P&L Matrix")
    ws_pnl.cell(row=1, column=1, value="P&L Matrix").font = Font(bold=True)
    ws_pnl.cell(row=3, column=1, value="% Change").font = Font(bold=True)
    ws_pnl.cell(row=3, column=2, value="Price").font = Font(bold=True)
    
    for col_idx, (dte, dt) in enumerate(date_columns, start=3):
        c = ws_pnl.cell(row=3, column=col_idx, value=dt)
        c.number_format = 'mmm d, yyyy'
        c.font = Font(bold=True)
    
    for r_idx, pct in enumerate(pct_steps, start=4):
        ws_pnl.cell(row=r_idx, column=1, value=pct).number_format = '0.00%'
        ws_pnl.cell(row=r_idx, column=2, value=f"='Option Price Matrix'!B{r_idx}").number_format = '$#,##0.00'
        for col_idx, _ in enumerate(date_columns, start=3):
            col_letter = get_column_letter(col_idx)
            opt_price_ref = f"'Option Price Matrix'!{col_letter}{r_idx}"
            entry_prem_ref = "'Assumptions'!$B$6"
            mult_ref = "'Assumptions'!$B$7"
            ws_pnl.cell(row=r_idx, column=col_idx, value=f"=({opt_price_ref}-{entry_prem_ref})*{mult_ref}").number_format = '$#,##0.00'

    # --- 4. Combined View ---
    ws_comb = wb.create_sheet("Combined View")
    ws_comb.cell(row=1, column=1, value="Combined View").font = Font(bold=True)
    ws_comb.cell(row=3, column=1, value="% Change").font = Font(bold=True)
    ws_comb.cell(row=3, column=2, value="Price").font = Font(bold=True)
    
    for col_idx, (dte, dt) in enumerate(date_columns, start=3):
        c = ws_comb.cell(row=3, column=col_idx, value=dt)
        c.number_format = 'mmm d, yyyy'
        c.font = Font(bold=True)
        ws_comb.column_dimensions[get_column_letter(col_idx)].width = 20

    for r_idx, pct in enumerate(pct_steps, start=4):
        ws_comb.cell(row=r_idx, column=1, value=pct).number_format = '0.00%'
        ws_comb.cell(row=r_idx, column=2, value=f"='Option Price Matrix'!B{r_idx}").number_format = '$#,##0.00'
        for col_idx, _ in enumerate(date_columns, start=3):
            col_letter = get_column_letter(col_idx)
            pnl_ref = f"'P&L Matrix'!{col_letter}{r_idx}"
            opt_ref = f"'Option Price Matrix'!{col_letter}{r_idx}"
            formula = f'=TEXT({opt_ref},"$0.00")&" / "&IF({pnl_ref}>0,"+$"&TEXT({pnl_ref},"#,##0"),IF({pnl_ref}<0,"-$"&TEXT(ABS({pnl_ref}),"#,##0"),"$0"))'
            ws_comb.cell(row=r_idx, column=col_idx, value=formula)

    return wb

